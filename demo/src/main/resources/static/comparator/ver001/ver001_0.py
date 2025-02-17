import sys
import json
import os
from datetime import datetime
from openpyxl import load_workbook

# 🔥 UTF-8 인코딩 강제 설정 (Windows 환경 대응)
sys.stdout.reconfigure(encoding="utf-8")

def get_sample_data(workbook):
    data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sample_data = [
            [convert_value(sheet.cell(row=i, column=j).value) for j in range(1, 31)]
            for i in range(1, 31)
        ]
        data[sheet_name] = sample_data
    return data

def convert_value(value):
    """ JSON 직렬화가 불가능한 데이터 타입을 변환 """
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")  # 날짜를 문자열로 변환
    if isinstance(value, str):
        return value.encode("utf-8", "ignore").decode("utf-8")  # 🔥 한글 인코딩 문제 해결
    return value

def main():
    if len(sys.argv) != 3:
        print(json.dumps({"error": "❌ Usage: python ver001_0.py <ARoot> <BRoot>"}))
        sys.exit(2)

    file_path_a = sys.argv[1]
    file_path_b = sys.argv[2]

    if not os.path.exists(file_path_a):
        print(json.dumps({"error": f"❌ File A not found: {file_path_a}"}))
        sys.exit(2)
    
    if not os.path.exists(file_path_b):
        print(json.dumps({"error": f"❌ File B not found: {file_path_b}"}))
        sys.exit(2)

    try:
        workbook_a = load_workbook(file_path_a, data_only=True)
        workbook_b = load_workbook(file_path_b, data_only=True)

        result = {
            "문화재정보A.xlsx": {
                "시트 수": len(workbook_a.sheetnames),
                "시트 이름": workbook_a.sheetnames,
                "예시 데이터": get_sample_data(workbook_a)
            },
            "문화재정보B.xlsx": {
                "시트 수": len(workbook_b.sheetnames),
                "시트 이름": workbook_b.sheetnames,
                "예시 데이터": get_sample_data(workbook_b)
            }
        }

        print(json.dumps(result, ensure_ascii=False, default=str))  # 🔥 UTF-8 보장

    except Exception as e:
        print(json.dumps({"error": f"❌ Python Error: {str(e)}"}))
        sys.exit(1)

if __name__ == "__main__":
    main()
